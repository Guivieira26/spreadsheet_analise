from openpyxl import load_workbook

#origem
plan_origin = load_workbook('Coldpack_paracetamol Clinical Trial.xlsx')
aba_origin = plan_origin.active

#destino
plan_dest = load_workbook('TOCOMPARE.xlsx')
aba_dest = plan_dest.active

y=2
linha = 2 #linha de destino
while y<=913:
    for col in range(1, 19):  # colunas de A (1) até Z (19)
        valor = aba_origin.cell(row=y, column=col).value
        aba_dest.cell(row=linha, column=col).value = valor
    y = y + 4
    linha = linha + 1
plan_dest.save('TOCOMPARE.xlsx')