from openpyxl import load_workbook
from pathlib import Path
def planilha_organization():
    base_dir = Path(__file__).resolve().parent.parent  # volta da /src/ pro diretório do projeto
    docs_dir = base_dir / "docs"

    plan_origin_path = docs_dir / "Coldpack_paracetamol Clinical Trial.xlsx"
    plan_dest_path = docs_dir / "new_tabel.xlsx"
    save_path = docs_dir / "new_tabel_editBy.xlsx"

    plan_origin = load_workbook(plan_origin_path)
    aba_origin = plan_origin.active

    plan_dest = load_workbook(plan_dest_path)
    aba_dest = plan_dest.active

    z = 2 #linha de destino
    y = 2 #linha de origem

    while y<=913: #sao 4 linhas para cada pessoa estudada e uma linha de instrução, (228*4) + 1 = 913
        # cada id passa 4 vezes para monitoramento de dor
        dor_ini = aba_origin[f'C{y}'].value
        y = y+3
        dor_end = aba_origin[f'C{y}'].value    
    #verificar se inicial é maior que ultimo para ver se a dor melhorou
    #Etapa 1:  se a dor melhorou
        if dor_ini is not None and dor_end is not None:
            if dor_ini > dor_end: 
                aba_dest[f'E{z}']= 'x'
    #Etapa 2: se fez uso de paracetamol
        valor = aba_origin[f'D{y}'].value
        if str(valor).lower() == "palacetamol": #compara evitando diferenças do tipo letras minusculas
            aba_dest[f'H{z}'] = 'x'
    #Etapa 3: se é jovem
        valor = aba_origin[f'E{y}'].value
        if valor <=19: 
            aba_dest[f'J{z}'] = 'x'
    #grupo adulto
        elif valor < 60:
            aba_dest[f'K{z}'] = 'x'
        #se for 60 ou mais é idoso
    #Etapa:4 se é casada
        valor = aba_origin[f'F{y}'].value
        if str(valor).lower() == "married":
            aba_dest[f'M{z}']= 'x'
        #se nao e solteiro
    #Etapa 5: se tem escolaridade
        valor = aba_origin[f'G{y}'].value
        if str(valor).lower() == "primary":
            aba_dest[f'N{z}'] = 'x'
        elif str(valor).lower() == "secondary":
            aba_dest[f'O{z}'] = 'x'
        #se nao nao tem escolaridade
    #Etapa 6: emprego
        valor = aba_origin[f'H{y}'].value
        if str(valor).lower() == "employed":
            aba_dest[f'P{z}'] = 'x'
        elif str(valor).strip().lower() == "housewife":
            aba_dest[f'Q{z}'] = 'x'
        #Senao e camponesa
    #Etapa 7: se e urbana
        valor = aba_origin[f'I{y}'].value
        if str(valor).lower() == "urban":
            aba_dest[f'S{z}'] = 'x'
        #else is rural
    #Etapa 8: visitas
        valor = aba_origin[f'j{y}'].value
        if valor <=2:
            aba_dest[f'T{z}']='x'
        elif valor<=5:
            aba_dest[f'U{z}']='x'
        #senao é alta
    #Etapa 9: quantidade de partos
        valor = aba_origin[f'K{y}'].value
        if valor == 1:
            print (valor)
            aba_dest[f'W{z}']='x'
            print ("etro mermo opaaaaaa")
            print ("coluna ",z,"linha ",y)
        elif valor == 2:
            aba_dest[f'X{z}']='x'
        elif valor == 3:
            aba_dest[f'Y{z}']='x'
        #senao e maior que 4
    #Etapa 10: renda
        valor = aba_origin[f'M{y}'].value    
        if str(valor).strip().lower() == "more than 1usd per day":        
            aba_dest[f'AA{z}']='x'
        #senao é menos que 1
    #Etapa 11: sexo
        valor = aba_origin[f'N{y}'].value
        if str(valor).lower() == "male":
            aba_dest[f'AB{z}'] = 'x'
        #senao e feminina
    #Etapa 12: se bebe tem mais de 3 kg
        valor = aba_origin[f'O{y}'].value
        if valor is not None and float(valor) > 3.0:
            aba_dest[f'AD{z}']='x'
        #senao e menor
    #Etapa 13: Se organismo intacto
        valor = aba_origin[f'P{y}'].value
        if str(valor).lower()=="intact":
            aba_dest[f'AE{z}']='x'
        #se nao ta rompido
    #Etapa 14: se amametou antes de 1 hora
        valor = aba_origin[f'R{y}'].value
        if str(valor).lower()!="after one hour":
            aba_dest[f'AF{z}']='x'
        #senao foi depois da primeira hora
        z=z+1
        y=y+1
    plan_dest.save(save_path)